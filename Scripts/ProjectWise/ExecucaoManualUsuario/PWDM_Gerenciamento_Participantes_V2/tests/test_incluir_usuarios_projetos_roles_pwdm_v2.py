import sys
import tempfile
import unittest
from pathlib import Path


PASTA_SCRIPT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PASTA_SCRIPT))

from incluir_usuarios_projetos_roles_pwdm_v2 import (  # noqa: E402
    LinhaInclusao,
    aplicar_operacao_rbac,
    buscar_convites_rbac,
    buscar_usuarios_rbac,
    carregar_linhas_inclusao_xlsx,
    extrair_id_usuario,
    extrair_role_ids_usuario,
    extrair_roles_de_membros,
    localizar_convite_por_email,
    localizar_usuario_por_email,
    normalizar_role,
    resolver_role,
    resolver_projetos_linha,
    valor_bool,
)


class PaginaRbacFake:
    def __init__(self, respostas):
        self.respostas = list(respostas)
        self.chamadas = []

    def wait_for_load_state(self, _state, timeout=None):
        return None

    def evaluate(self, _script, argumentos):
        self.chamadas.append(argumentos)
        return self.respostas.pop(0)


class IncluirUsuariosProjetosRolesTests(unittest.TestCase):
    def test_buscar_usuarios_rbac_percorre_todas_as_paginas(self):
        pagina = PaginaRbacFake(
            [
                {
                    "ok": True,
                    "status": 200,
                    "body": {
                        "members": [{"email": f"usuario{i}@empresa.com"} for i in range(100)],
                        "_links": {"next": {"href": "pagina-2"}},
                    },
                },
                {
                    "ok": True,
                    "status": 200,
                    "body": {"members": [{"email": "ultimo@empresa.com"}], "_links": {}},
                },
            ]
        )

        usuarios = buscar_usuarios_rbac(pagina, "itwin-1")

        self.assertEqual(len(usuarios), 101)
        self.assertIn("$top=100&$skip=0", pagina.chamadas[0]["path"])
        self.assertIn("$top=100&$skip=100", pagina.chamadas[1]["path"])

    def test_buscar_convites_e_localizar_email_normalizado(self):
        pagina = PaginaRbacFake(
            [
                {
                    "ok": True,
                    "status": 200,
                    "body": {
                        "invitations": [
                            {"id": "convite-1", "email": " Usuario@Empresa.COM ", "roles": []}
                        ],
                        "_links": {},
                    },
                }
            ]
        )

        convites = buscar_convites_rbac(pagina, "itwin-1")

        self.assertEqual(localizar_convite_por_email(convites, "usuario@empresa.com")["id"], "convite-1")
        self.assertIn("/members/invitations?", pagina.chamadas[0]["path"])

    def test_localizar_usuario_normaliza_espacos_e_caixa(self):
        usuarios = [{"id": "user-1", "email": " Usuario@Empresa.COM "}]

        usuario = localizar_usuario_por_email(usuarios, "usuario@empresa.com")

        self.assertEqual(usuario["id"], "user-1")

    def test_team_member_exists_reconsulta_e_atualiza_role(self):
        pagina = PaginaRbacFake(
            [
                {
                    "ok": False,
                    "status": 409,
                    "body": {"error": {"code": "TeamMemberExists"}},
                },
                {
                    "ok": True,
                    "status": 200,
                    "body": {
                        "members": [
                            {
                                "id": "user-1",
                                "email": "usuario@empresa.com",
                                "roles": [{"id": "role-antiga"}],
                            }
                        ],
                        "_links": {},
                    },
                },
                {"ok": True, "status": 200, "body": {"member": {"id": "user-1"}}},
            ]
        )
        operacao = {
            "acaoEfetiva": "adicionar_usuario",
            "projectId": "itwin-1",
            "email": "usuario@empresa.com",
            "roleResolvida": {"id": "role-nova"},
        }

        resultado = aplicar_operacao_rbac(pagina, operacao)

        self.assertEqual(resultado["status"], "roles_atualizadas_apos_conflito")
        self.assertEqual(pagina.chamadas[2]["method"], "PATCH")
        self.assertEqual(pagina.chamadas[2]["payload"]["roleIds"], ["role-antiga", "role-nova"])

    def test_team_member_exists_com_convite_e_role_vira_sem_alteracao(self):
        pagina = PaginaRbacFake(
            [
                {"ok": False, "status": 409, "body": {"error": {"code": "TeamMemberExists"}}},
                {"ok": True, "status": 200, "body": {"members": [], "_links": {}}},
                {
                    "ok": True,
                    "status": 200,
                    "body": {
                        "invitations": [
                            {
                                "id": "convite-1",
                                "email": "usuario@empresa.com",
                                "roles": [{"id": "role-nova"}],
                            }
                        ],
                        "_links": {},
                    },
                },
            ]
        )
        operacao = {
            "acaoEfetiva": "adicionar_usuario",
            "projectId": "itwin-1",
            "email": "usuario@empresa.com",
            "roleResolvida": {"id": "role-nova"},
        }

        resultado = aplicar_operacao_rbac(pagina, operacao)

        self.assertEqual(resultado["status"], "sem_alteracao")

    def test_carregar_linhas_inclusao_xlsx_com_role(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as pasta:
            caminho = Path(pasta) / "inclusoes.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["email", "projeto", "role"])
            sheet.append(["Usuario1@Empresa.com", "Projeto A", "PWDM Issuer"])
            sheet.append(["sem-email", "Projeto A", "Team Member"])
            workbook.save(caminho)

            resultado = carregar_linhas_inclusao_xlsx(caminho)

        linhas = resultado["linhas"]
        self.assertEqual(len(linhas), 1)
        self.assertEqual(linhas[0].email, "usuario1@empresa.com")
        self.assertEqual(linhas[0].projeto, "Projeto A")
        self.assertEqual(linhas[0].role, "PWDM Issuer")
        self.assertEqual(resultado["invalidas"][0]["motivo"], "email_invalido")

    def test_carregar_linhas_inclusao_xlsx_sem_role(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as pasta:
            caminho = Path(pasta) / "inclusoes.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["email", "projeto"])
            sheet.append(["Usuario1@Empresa.com", "Projeto A"])
            workbook.save(caminho)

            resultado = carregar_linhas_inclusao_xlsx(caminho)

        linhas = resultado["linhas"]
        self.assertEqual(len(linhas), 1)
        self.assertEqual(linhas[0].email, "usuario1@empresa.com")
        self.assertEqual(linhas[0].projeto, "Projeto A")
        self.assertEqual(linhas[0].role, "")

    def test_carregar_linhas_inclusao_xlsx_apenas_com_coluna_email(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as pasta:
            caminho = Path(pasta) / "usuarios.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["E-mail"])
            sheet.append(["Usuario1@Empresa.com"])
            sheet.append(["usuario2@empresa.com"])
            workbook.save(caminho)

            resultado = carregar_linhas_inclusao_xlsx(caminho)

        linhas = resultado["linhas"]
        self.assertEqual([linha.email for linha in linhas], ["usuario1@empresa.com", "usuario2@empresa.com"])
        self.assertEqual([linha.projeto for linha in linhas], ["todos", "todos"])

    def test_resolver_projetos_linha_por_projectwise_id(self):
        projetos = [
            {
                "nome": "Projeto A",
                "projectId": "11111111-1111-1111-1111-111111111111",
                "connectSpaceId": "11111111-1111-1111-1111-111111111111",
                "origemProjectWise": {"id": "123", "nome": "Projeto A"},
            }
        ]
        linha = LinhaInclusao(
            linha=2,
            email="usuario@empresa.com",
            projeto="123",
            role="Team Member",
        )

        encontrados, criterio = resolver_projetos_linha(linha, projetos)

        self.assertEqual(criterio, "exato")
        self.assertEqual(encontrados, projetos)

    def test_resolver_projetos_linha_todos(self):
        projetos = [
            {"nome": "Projeto A", "projectId": "1", "connectSpaceId": "1", "origemProjectWise": {}},
            {"nome": "Projeto B", "projectId": "2", "connectSpaceId": "2", "origemProjectWise": {}},
        ]
        linha = LinhaInclusao(
            linha=2,
            email="usuario@empresa.com",
            projeto="todos",
            role="Team Member",
        )

        encontrados, criterio = resolver_projetos_linha(linha, projetos)

        self.assertEqual(criterio, "todos")
        self.assertEqual(encontrados, projetos)

    def test_valor_bool_aceita_s_n_e_x(self):
        self.assertEqual(valor_bool("S"), True)
        self.assertEqual(valor_bool("x"), True)
        self.assertEqual(valor_bool("nao"), False)
        self.assertIsNone(valor_bool("talvez"))

    def test_resolver_role_por_nome_exato(self):
        roles = [
            {"id": "1", "displayName": "Team Member"},
            {"id": "2", "displayName": "PWDM Issuer"},
        ]

        role = resolver_role("pwdm issuer", roles)

        self.assertEqual(role["id"], "2")

    def test_normalizar_role_aceita_role_id(self):
        role = normalizar_role({"roleId": "abc", "name": "Team Member"}, "iTwin", "itwin-1")

        self.assertEqual(role["id"], "abc")
        self.assertEqual(role["displayName"], "Team Member")

    def test_extrair_ids_usuario_em_formatos_diferentes(self):
        usuario = {
            "userId": "user-1",
            "roleIds": ["role-1"],
            "roles": [{"roleId": "role-2"}, {"id": "role-3"}],
        }

        self.assertEqual(extrair_id_usuario(usuario), "user-1")
        self.assertEqual(extrair_role_ids_usuario(usuario), ["role-1", "role-2", "role-3"])

    def test_extrair_roles_de_membros(self):
        usuarios = [
            {
                "email": "usuario@empresa.com",
                "roles": [
                    {"id": "team-member-id", "displayName": "Team Member"},
                    {"roleId": "imodel-id", "name": "iModel", "type": "enterprise"},
                ],
            }
        ]

        roles = extrair_roles_de_membros(usuarios, "itwin-1")

        self.assertEqual([role["displayName"] for role in roles], ["iModel", "Team Member"])
        self.assertEqual(roles[0]["type"], "enterprise")


if __name__ == "__main__":
    unittest.main()
