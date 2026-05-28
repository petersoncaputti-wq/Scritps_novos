from pathlib import Path
from typing import Any

from regras_v2 import EMAIL_REGEX


CABECALHOS_EMAIL = {"email", "e-mail", "e mail"}


def normalizar_email(valor: Any) -> str:
    return str(valor or "").strip().lower()


def _normalizar_cabecalho(valor: Any) -> str:
    return " ".join(str(valor or "").strip().lower().replace("-", " ").split())


def localizar_coluna_email(cabecalhos: list[Any]) -> int:
    for indice, cabecalho in enumerate(cabecalhos):
        if _normalizar_cabecalho(cabecalho) in CABECALHOS_EMAIL:
            return indice
    raise ValueError("A planilha deve ter uma coluna chamada 'email'.")


def deduplicar_emails(emails: list[str]) -> tuple[list[str], list[str]]:
    unicos: list[str] = []
    duplicados: list[str] = []
    vistos: set[str] = set()

    for email in emails:
        email_normalizado = normalizar_email(email)
        if not email_normalizado:
            continue
        if email_normalizado in vistos:
            duplicados.append(email_normalizado)
            continue
        vistos.add(email_normalizado)
        unicos.append(email_normalizado)

    return unicos, duplicados


def carregar_emails_xlsx(caminho: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as erro:
        raise RuntimeError("Dependencia ausente: instale o pacote openpyxl para ler arquivos .xlsx.") from erro

    caminho = Path(caminho)
    if caminho.suffix.lower() != ".xlsx":
        raise ValueError("Selecione uma planilha no formato .xlsx.")
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {caminho}")

    workbook = load_workbook(caminho, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        linhas = sheet.iter_rows(values_only=True)
        try:
            cabecalhos = list(next(linhas))
        except StopIteration as erro:
            raise ValueError("A planilha esta vazia.") from erro

        coluna_email = localizar_coluna_email(cabecalhos)
        encontrados: list[str] = []
        invalidos: list[dict[str, Any]] = []
        linhas_sem_email: list[int] = []
        total_linhas_dados = 0

        for numero_linha, linha in enumerate(linhas, start=2):
            total_linhas_dados += 1
            valor = linha[coluna_email] if coluna_email < len(linha) else ""
            email = normalizar_email(valor)
            if not email:
                linhas_sem_email.append(numero_linha)
                continue
            if not EMAIL_REGEX.match(email):
                invalidos.append({"linha": numero_linha, "email": email})
                continue
            encontrados.append(email)

        unicos, duplicados = deduplicar_emails(encontrados)
        if not unicos:
            raise ValueError("Nenhum e-mail valido foi encontrado na planilha.")

        return {
            "arquivo": str(caminho),
            "aba": sheet.title,
            "linhasDados": total_linhas_dados,
            "emailsEncontrados": len(encontrados),
            "emailsValidosUnicos": unicos,
            "duplicados": duplicados,
            "invalidos": invalidos,
            "linhasSemEmail": linhas_sem_email,
        }
    finally:
        workbook.close()
