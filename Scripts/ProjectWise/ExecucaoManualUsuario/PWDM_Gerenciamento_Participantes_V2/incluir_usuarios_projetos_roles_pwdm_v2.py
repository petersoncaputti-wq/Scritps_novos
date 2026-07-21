import json
import time
import traceback
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tkinter import Tk, filedialog
from typing import Any, Optional

from playwright.sync_api import Browser, Page, sync_playwright

from gerenciar_participante_pwdm_v2 import (
    EMAIL_REGEX,
    iniciar_navegador,
    preparar_pasta_logs,
    salvar_log,
    solicitar_email,
)
from gerenciar_participante_pwdm_connected_v2 import (
    carregar_projetos_para_fluxo,
    executar_seletor_projectwise,
)
from regras_v2 import normalizar_texto_chave


PASTA_BASE = Path(__file__).resolve().parent
PASTA_LOGS = PASTA_BASE / "Logs"
URL_RBAC_PORTAL = "https://connect-rbacportal.bentley.com"
URL_RBAC_API = "https://api.bentley.com/accesscontrol/itwins"
URL_ITWINS_API = "https://api.bentley.com/itwins"

CABECALHOS = {
    "email": {"email", "e-mail", "e mail", "usuario", "usuário", "user"},
    "projeto": {
        "projeto",
        "nome projeto",
        "nome do projeto",
        "project",
        "project name",
        "projectid",
        "project id",
        "connectedprojectid",
        "connected project id",
        "pw id",
        "projectwise id",
    },
    "role": {"role", "roles", "perfil", "titulo", "título", "cargo", "role title", "roletitle"},
}


@dataclass(frozen=True)
class LinhaInclusao:
    linha: int
    email: str
    projeto: str
    role: str


def nome_execucao() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def normalizar_cabecalho(valor: Any) -> str:
    return normalizar_texto_chave(str(valor or "").replace("_", " "))


def normalizar_email(valor: Any) -> str:
    return str(valor or "").strip().lower()


def valor_bool(valor: Any) -> Optional[bool]:
    if valor is None:
        return None
    if isinstance(valor, bool):
        return valor
    if isinstance(valor, (int, float)) and valor in {0, 1}:
        return bool(valor)

    texto = normalizar_texto_chave(str(valor))
    if texto in {"s", "sim", "y", "yes", "1", "true", "verdadeiro", "x"}:
        return True
    if texto in {"n", "nao", "no", "0", "false", "falso", ""}:
        return False
    return None


def localizar_colunas(cabecalhos: list[Any]) -> dict[str, int]:
    colunas: dict[str, int] = {}
    cabecalhos_norm = [normalizar_cabecalho(cabecalho) for cabecalho in cabecalhos]

    for campo, aliases in CABECALHOS.items():
        aliases_norm = {normalizar_cabecalho(alias) for alias in aliases}
        for indice, cabecalho in enumerate(cabecalhos_norm):
            if cabecalho in aliases_norm:
                colunas[campo] = indice
                break

    if "email" not in colunas:
        raise ValueError("A planilha deve ter uma coluna de e-mail.")

    return colunas


def valor_linha(linha: tuple[Any, ...], indice: Optional[int]) -> Any:
    if indice is None or indice >= len(linha):
        return ""
    return linha[indice]


def carregar_linhas_inclusao_xlsx(caminho: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as erro:
        raise RuntimeError("Dependencia ausente: instale o pacote openpyxl para ler arquivos .xlsx.") from erro

    caminho = Path(caminho)
    if caminho.suffix.lower() != ".xlsx":
        raise ValueError("Selecione uma planilha no formato .xlsx.")
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {caminho}")

    workbook = load_workbook(caminho, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        linhas_iter = sheet.iter_rows(values_only=True)
        try:
            cabecalhos = list(next(linhas_iter))
        except StopIteration as erro:
            raise ValueError("A planilha esta vazia.") from erro

        colunas = localizar_colunas(cabecalhos)
        tem_coluna_projeto = "projeto" in colunas
        linhas_validas: list[LinhaInclusao] = []
        invalidas: list[dict[str, Any]] = []
        linhas_vazias = 0

        for numero_linha, linha in enumerate(linhas_iter, start=2):
            email = normalizar_email(valor_linha(linha, colunas.get("email")))
            projeto_informado = str(valor_linha(linha, colunas.get("projeto")) or "").strip()
            projeto = projeto_informado if tem_coluna_projeto else "todos"
            role = str(valor_linha(linha, colunas.get("role")) or "").strip()

            if not email and not projeto_informado and not role:
                linhas_vazias += 1
                continue
            if not EMAIL_REGEX.match(email):
                invalidas.append({"linha": numero_linha, "motivo": "email_invalido", "email": email})
                continue
            if tem_coluna_projeto and not projeto:
                invalidas.append({"linha": numero_linha, "motivo": "projeto_vazio", "email": email})
                continue

            linhas_validas.append(LinhaInclusao(numero_linha, email, projeto, role))

        if not linhas_validas:
            raise ValueError("Nenhuma linha valida foi encontrada na planilha.")

        return {
            "arquivo": str(caminho),
            "aba": sheet.title,
            "linhas": linhas_validas,
            "invalidas": invalidas,
            "linhasVazias": linhas_vazias,
        }
    finally:
        workbook.close()


def selecionar_planilha_xlsx() -> Path:
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        caminho = filedialog.askopenfilename(
            title="Selecione a planilha de inclusao de usuarios por projeto",
            filetypes=[("Planilhas Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
        )
    finally:
        root.destroy()

    if not caminho:
        raise RuntimeError("Nenhuma planilha foi selecionada.")
    return Path(caminho)


def solicitar_modo_usuarios() -> str:
    print("\nModo de usuarios")
    print("01. Usuario unico")
    print("02. Lote por planilha .xlsx")
    while True:
        resposta = input("Escolha [1/2]: ").strip()
        if resposta in {"1", "01"}:
            return "unico"
        if resposta in {"2", "02"}:
            return "lote_xlsx"
        print("[AVISO] Escolha 1 ou 2.")


def confirmar_email_informado(email: str) -> None:
    print(f"\nE-mail informado: {email}")
    while True:
        resposta = input("Continuar com este e-mail? [S/N]: ").strip().lower()
        if resposta in {"s", "sim", "y", "yes"}:
            return
        if resposta in {"n", "nao", "não", "no", ""}:
            raise RuntimeError("E-mail nao confirmado. Nenhuma alteracao sera preparada.")
        print("[AVISO] Responda apenas S ou N.")


def carregar_linhas_usuario_unico() -> list[LinhaInclusao]:
    email = solicitar_email()
    confirmar_email_informado(email)
    return [LinhaInclusao(linha=0, email=email, projeto="todos", role="")]


def carregar_linhas_planilha_interativa() -> list[LinhaInclusao]:
    caminho_planilha = selecionar_planilha_xlsx()
    resumo = carregar_linhas_inclusao_xlsx(caminho_planilha)
    linhas = list(resumo["linhas"])
    print(f"\n[OK] Planilha importada: {Path(resumo['arquivo']).resolve()}")
    print(f"- Aba: {resumo['aba']}")
    print(f"- Linhas validas: {len(linhas)}")
    print(f"- Linhas invalidas ignoradas: {len(resumo['invalidas'])}")
    print("- Projeto: sem coluna de projeto, os usuarios serao aplicados a todos os projetos selecionados.")
    print("- Role: sera selecionada no RBAC para as linhas sem coluna/valor de role.")
    if resumo["invalidas"]:
        for item in resumo["invalidas"][:20]:
            print(f"  - Linha {item['linha']}: {item['motivo']} ({item.get('email') or ''})")
    return linhas


def chaves_projeto(projeto: dict[str, Any]) -> set[str]:
    origem = projeto.get("origemProjectWise") or {}
    valores = {
        projeto.get("projectId"),
        projeto.get("connectSpaceId"),
        projeto.get("nome"),
        projeto.get("numero"),
        origem.get("id"),
        origem.get("guid"),
        origem.get("connectedProjectId"),
        origem.get("nome"),
        origem.get("projectWiseWebName"),
        origem.get("projectWiseWebNumber"),
    }
    return {normalizar_texto_chave(str(valor)) for valor in valores if str(valor or "").strip()}


def resolver_projetos_linha(linha: LinhaInclusao, projetos: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], str]:
    projeto_chave = normalizar_texto_chave(linha.projeto)
    if projeto_chave in {"todos", "todo", "all", "*"}:
        return projetos, "todos"

    correspondentes = [projeto for projeto in projetos if projeto_chave in chaves_projeto(projeto)]
    if correspondentes:
        return correspondentes, "exato"

    parciais = [
        projeto
        for projeto in projetos
        if any(projeto_chave and projeto_chave in chave for chave in chaves_projeto(projeto))
    ]
    if len(parciais) == 1:
        return parciais, "parcial"
    if len(parciais) > 1:
        nomes = ", ".join(str(item.get("nome") or item.get("projectId")) for item in parciais[:5])
        raise ValueError(f"Linha {linha.linha}: projeto ambiguo '{linha.projeto}'. Correspondencias: {nomes}")

    raise ValueError(f"Linha {linha.linha}: projeto nao encontrado na selecao: {linha.projeto}")


def url_tela_membros_rbac(itwin_id: str) -> str:
    return f"{URL_RBAC_PORTAL}/Manage/{itwin_id}/users"


def abrir_portal_rbac(page: Page, itwin_id: str) -> None:
    url = url_tela_membros_rbac(itwin_id)
    print(f"[2/9] Acessando tela de membros RBAC: {url}")
    page.goto(url, wait_until="domcontentloaded")
    print("\nSe necessario, faca login manualmente no navegador.")
    input("Quando estiver autenticado e a tela carregar, pressione ENTER no terminal...")


def rbac_api(page: Page, method: str, path: str, payload: Optional[dict[str, Any]] = None) -> dict[str, Any]:
    ultimo_erro: Optional[Exception] = None
    for tentativa in range(3):
        try:
            page.wait_for_load_state("domcontentloaded", timeout=30_000)
            return page.evaluate(
        """
        async ({ method, path, payload, apiBase }) => {
            const tokenResponse = await fetch("https://connect-rbacportal.bentley.com/Manage/GetAccessToken", {
                method: "GET",
                credentials: "include",
                headers: { "Accept": "application/json" }
            });
            if (!tokenResponse.ok) {
                const text = await tokenResponse.text();
                return { ok: false, status: tokenResponse.status, statusText: tokenResponse.statusText, text };
            }
            const tokenBody = await tokenResponse.json();
            const accessToken = tokenBody.AccessToken || tokenBody.accessToken || tokenBody.token || "";
            const response = await fetch(apiBase + path, {
                method,
                headers: {
                    "Authorization": `Bearer ${accessToken}`,
                    "Accept": "application/vnd.bentley.itwin-platform.v2+json",
                    "Content-Type": "application/json"
                },
                body: payload ? JSON.stringify(payload) : undefined
            });
            const text = await response.text();
            let body = null;
            try {
                body = text ? JSON.parse(text) : null;
            } catch {
                body = text;
            }
            return {
                ok: response.ok,
                status: response.status,
                statusText: response.statusText,
                body,
                text
            };
        }
        """,
                {"method": method, "path": path, "payload": payload, "apiBase": URL_RBAC_API},
            )
        except Exception as erro:
            ultimo_erro = erro
            if "Execution context was destroyed" not in str(erro) or tentativa == 2:
                raise
            print("  [AVISO] A pagina ainda estava navegando; repetindo a consulta RBAC...")
            time.sleep(1)

    raise RuntimeError(f"Falha ao consultar RBAC: {ultimo_erro}")


def bentley_api(
    page: Page,
    method: str,
    api_base: str,
    path: str,
    payload: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    return page.evaluate(
        """
        async ({ method, path, payload, apiBase }) => {
            const tokenResponse = await fetch("https://connect-rbacportal.bentley.com/Manage/GetPrivateToken", {
                method: "GET",
                credentials: "include",
                headers: { "Accept": "application/json" }
            });
            if (!tokenResponse.ok) {
                const text = await tokenResponse.text();
                return { ok: false, status: tokenResponse.status, statusText: tokenResponse.statusText, text };
            }
            const tokenText = await tokenResponse.text();
            let tokenBody = null;
            try {
                tokenBody = tokenText ? JSON.parse(tokenText) : null;
            } catch {
                tokenBody = tokenText;
            }
            let accessToken = "";
            if (typeof tokenBody === "string") {
                accessToken = tokenBody;
            } else if (tokenBody) {
                accessToken = tokenBody.AccessToken || tokenBody.accessToken || tokenBody.PrivateToken || tokenBody.privateToken || tokenBody.token || "";
            }
            const authorization = accessToken.toLowerCase().startsWith("bearer ")
                ? accessToken
                : `Bearer ${accessToken}`;
            const response = await fetch(apiBase + path, {
                method,
                headers: {
                    "Authorization": authorization,
                    "Accept": "application/vnd.bentley.itwin-platform.v2+json",
                    "Content-Type": "application/json"
                },
                body: payload ? JSON.stringify(payload) : undefined
            });
            const text = await response.text();
            let body = null;
            try {
                body = text ? JSON.parse(text) : null;
            } catch {
                body = text;
            }
            return {
                ok: response.ok,
                status: response.status,
                statusText: response.statusText,
                body,
                text
            };
        }
        """,
        {"method": method, "path": path, "payload": payload, "apiBase": api_base},
    )


def exigir_rbac_ok(resultado: dict[str, Any], contexto: str) -> Any:
    if resultado.get("ok"):
        return resultado.get("body")
    detalhe = resultado.get("body") or resultado.get("text") or ""
    raise RuntimeError(
        f"{contexto}: HTTP {resultado.get('status')} {resultado.get('statusText')} - {detalhe}"
    )


def lista_de_resposta(body: Any, chave: str) -> list[dict[str, Any]]:
    if isinstance(body, dict) and isinstance(body.get(chave), list):
        return [item for item in body[chave] if isinstance(item, dict)]
    if isinstance(body, dict) and isinstance(body.get("items"), list):
        return [item for item in body["items"] if isinstance(item, dict)]
    if isinstance(body, list):
        return [item for item in body if isinstance(item, dict)]
    return []


def normalizar_role(role: dict[str, Any], tipo: str, itwin_id: str) -> dict[str, Any]:
    return {
        "id": str(role.get("id") or role.get("roleId") or ""),
        "displayName": str(role.get("displayName") or role.get("name") or ""),
        "description": str(role.get("description") or ""),
        "permissions": role.get("permissions") or [],
        "type": tipo,
        "iTwinId": itwin_id,
    }


def extrair_primeiro_id(body: Any) -> str:
    if not isinstance(body, dict):
        return ""

    candidatos = [
        body.get("id"),
        body.get("iTwinId"),
        body.get("itwinId"),
        body.get("parentId"),
        body.get("parentITwinId"),
    ]
    for chave in ("iTwin", "itwin", "parent", "parentITwin", "account"):
        valor = body.get(chave)
        if isinstance(valor, dict):
            candidatos.extend(
                [
                    valor.get("id"),
                    valor.get("iTwinId"),
                    valor.get("itwinId"),
                    valor.get("parentId"),
                    valor.get("parentITwinId"),
                ]
            )

    for candidato in candidatos:
        texto = str(candidato or "").strip()
        if texto:
            return texto
    return ""


def extrair_id_conta_itwin(body: Any, itwin_id: str) -> str:
    if not isinstance(body, dict):
        return ""

    candidatos = [
        body.get("accountId"),
        body.get("accountOwnerId"),
        body.get("iTwinAccountId"),
        body.get("parentId"),
        body.get("parentITwinId"),
    ]

    for chave in ("account", "iTwinAccount", "parent", "parentITwin", "iTwin", "itwin"):
        valor = body.get(chave)
        if isinstance(valor, dict):
            candidatos.extend(
                [
                    valor.get("id"),
                    valor.get("accountId"),
                    valor.get("accountOwnerId"),
                    valor.get("iTwinAccountId"),
                    valor.get("parentId"),
                    valor.get("parentITwinId"),
                ]
            )

    for candidato in candidatos:
        texto = str(candidato or "").strip()
        if texto and texto != itwin_id:
            return texto
    return ""


def buscar_itwin_pai(page: Page, itwin_id: str) -> str:
    tentativas = [
        (URL_ITWINS_API, f"/{itwin_id}/account", "Buscar conta do iTwin"),
        (URL_ITWINS_API, f"/{itwin_id}", "Buscar dados do iTwin"),
        (URL_ITWINS_API, "/myprimaryaccount", "Buscar conta primaria"),
    ]

    erros: list[str] = []
    for api_base, path, contexto in tentativas:
        try:
            body = exigir_rbac_ok(bentley_api(page, "GET", api_base, path), f"{contexto} {itwin_id}")
            parent_id = extrair_id_conta_itwin(body, itwin_id) or extrair_primeiro_id(body)
            if parent_id and parent_id != itwin_id:
                print(f"  [INFO] iTwin pai/conta identificado para roles enterprise: {parent_id}")
                return parent_id
        except Exception as erro:
            erros.append(str(erro))

    if erros:
        print("  [AVISO] Nao foi possivel descobrir iTwin pai/conta para roles enterprise.")
        for erro in erros[:3]:
            print(f"    - {erro}")
    return ""


def deduplicar_roles(roles: list[dict[str, Any]]) -> list[dict[str, Any]]:
    deduplicadas: dict[str, dict[str, Any]] = {}
    for role in roles:
        role_id = role.get("id")
        if not role_id:
            continue
        atual = deduplicadas.get(role_id)
        if not atual or atual.get("type") != "enterprise":
            deduplicadas[role_id] = role

    resultado = list(deduplicadas.values())
    resultado.sort(key=lambda item: (item["type"] != "enterprise", normalizar_texto_chave(item["displayName"])))
    return resultado


def extrair_roles_de_membros(usuarios: list[dict[str, Any]], itwin_id: str) -> list[dict[str, Any]]:
    roles: list[dict[str, Any]] = []
    for usuario in usuarios:
        for role in usuario.get("roles") or []:
            if not isinstance(role, dict):
                continue
            tipo = str(role.get("type") or "").strip() or "enterprise"
            roles.append(normalizar_role(role, tipo, itwin_id))
    return deduplicar_roles([role for role in roles if role["id"] and role["displayName"]])


def buscar_roles_rbac(page: Page, itwin_id: str) -> list[dict[str, Any]]:
    body_itwin = exigir_rbac_ok(rbac_api(page, "GET", f"/{itwin_id}/roles"), f"Buscar roles do iTwin {itwin_id}")
    roles = [normalizar_role(role, "iTwin", itwin_id) for role in lista_de_resposta(body_itwin, "roles")]

    roles = [role for role in roles if role["id"] and role["displayName"]]
    return deduplicar_roles(roles)


def mapa_roles_por_nome(roles: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    mapa: dict[str, dict[str, Any]] = {}
    for role in roles:
        chave = normalizar_texto_chave(role["displayName"])
        if chave and chave not in mapa:
            mapa[chave] = role
    return mapa


def resolver_role(role_informada: str, roles: list[dict[str, Any]]) -> dict[str, Any]:
    chave = normalizar_texto_chave(role_informada)
    mapa = mapa_roles_por_nome(roles)
    if chave in mapa:
        return mapa[chave]

    parciais = [role for role in roles if chave and chave in normalizar_texto_chave(role["displayName"])]
    if len(parciais) == 1:
        return parciais[0]
    if len(parciais) > 1:
        nomes = ", ".join(role["displayName"] for role in parciais[:8])
        raise ValueError(f"Role ambigua '{role_informada}'. Correspondencias: {nomes}")

    disponiveis = ", ".join(role["displayName"] for role in roles[:20])
    raise ValueError(f"Role nao encontrada: {role_informada}. Roles disponiveis: {disponiveis}")


def selecionar_role_padrao_por_lista(roles: list[dict[str, Any]]) -> str:
    if not roles:
        raise RuntimeError("Nenhuma role foi retornada pelo RBAC para selecao.")

    indice_padrao: Optional[int] = None
    for indice, role in enumerate(roles):
        if normalizar_texto_chave(role["displayName"]) == "team member":
            indice_padrao = indice
            break

    print("\nRole RBAC para linhas sem role informada")
    if indice_padrao is not None:
        print("Escolha pelo numero abaixo. ENTER usa Team Member.")
    else:
        print("Escolha pelo numero abaixo. ENTER nao sera aceito porque Team Member nao foi retornado pela API.")
    for indice, role in enumerate(roles, start=1):
        marcador = " [padrao]" if indice_padrao is not None and indice - 1 == indice_padrao else ""
        print(f"{indice:02d}. {role['displayName']} ({role['type']}){marcador}")

    while True:
        resposta = input("Numero ou nome da role: ").strip()
        if not resposta:
            if indice_padrao is not None:
                return roles[indice_padrao]["displayName"]
            print("[AVISO] Informe o numero ou nome exato da role desejada.")
            continue
        if resposta.isdigit():
            indice = int(resposta)
            if 1 <= indice <= len(roles):
                return roles[indice - 1]["displayName"]
            print(f"[AVISO] Informe um numero entre 1 e {len(roles)}.")
            continue

        try:
            return resolver_role(resposta, roles)["displayName"]
        except Exception as erro:
            print(f"[AVISO] {erro}")


def preencher_roles_pendentes(page: Page, projeto_referencia_id: str, linhas: list[LinhaInclusao]) -> list[LinhaInclusao]:
    if not any(not linha.role for linha in linhas):
        return linhas

    roles = buscar_roles_rbac(page, projeto_referencia_id)
    try:
        usuarios = buscar_usuarios_rbac(page, projeto_referencia_id)
        roles_membros = extrair_roles_de_membros(usuarios, projeto_referencia_id)
        if roles_membros:
            print(f"  [INFO] Roles extraidas dos membros existentes: {len(roles_membros)}")
            roles = deduplicar_roles(roles + roles_membros)
    except Exception as erro:
        print(f"  [AVISO] Nao foi possivel complementar roles pelos membros existentes: {erro}")

    role_padrao = selecionar_role_padrao_por_lista(roles)
    print(f"[OK] Role selecionada para linhas sem role: {role_padrao}")
    return [
        linha if linha.role else LinhaInclusao(linha.linha, linha.email, linha.projeto, role_padrao)
        for linha in linhas
    ]


def buscar_usuarios_rbac(page: Page, itwin_id: str) -> list[dict[str, Any]]:
    body = exigir_rbac_ok(rbac_api(page, "GET", f"/{itwin_id}/members/users"), f"Buscar usuarios {itwin_id}")
    return lista_de_resposta(body, "members") or lista_de_resposta(body, "users")


def localizar_usuario_por_email(usuarios: list[dict[str, Any]], email: str) -> Optional[dict[str, Any]]:
    email_chave = email.lower()
    for usuario in usuarios:
        if str(usuario.get("email") or "").lower() == email_chave:
            return usuario
    return None


def extrair_role_ids_usuario(usuario: dict[str, Any]) -> list[str]:
    role_ids: list[str] = []

    for campo in ("roleIds", "rolesIds"):
        valores = usuario.get(campo)
        if isinstance(valores, list):
            role_ids.extend(str(valor) for valor in valores if valor)

    roles = usuario.get("roles")
    if isinstance(roles, list):
        for role in roles:
            if isinstance(role, dict):
                role_id = role.get("id") or role.get("roleId")
                if role_id:
                    role_ids.append(str(role_id))
            elif role:
                role_ids.append(str(role))

    return list(dict.fromkeys(role_ids))


def extrair_id_usuario(usuario: dict[str, Any]) -> str:
    return str(usuario.get("id") or usuario.get("userId") or usuario.get("memberId") or "")


def montar_operacoes_rbac(page: Page, projetos: list[dict[str, Any]], linhas: list[LinhaInclusao]) -> list[dict[str, Any]]:
    print("\n[7/9] Lendo roles e membros nos projetos selecionados...")
    operacoes: list[dict[str, Any]] = []
    cache_roles: dict[str, list[dict[str, Any]]] = {}
    cache_usuarios: dict[str, list[dict[str, Any]]] = {}

    for indice, linha in enumerate(linhas, start=1):
        projetos_linha, criterio = resolver_projetos_linha(linha, projetos)
        print(
            f"- Linha {linha.linha or '-'} ({indice}/{len(linhas)}): "
            f"{linha.email} -> {len(projetos_linha)} projeto(s) [{criterio}]"
        )

        for projeto in projetos_linha:
            itwin_id = projeto["projectId"]
            try:
                if itwin_id not in cache_roles:
                    print(f"  Consultando roles: {projeto['nome']} ({itwin_id})")
                    cache_roles[itwin_id] = buscar_roles_rbac(page, itwin_id)
                if itwin_id not in cache_usuarios:
                    print(f"  Consultando membros: {projeto['nome']} ({itwin_id})")
                    cache_usuarios[itwin_id] = buscar_usuarios_rbac(page, itwin_id)
                    roles_membros = extrair_roles_de_membros(cache_usuarios[itwin_id], itwin_id)
                    if roles_membros:
                        cache_roles[itwin_id] = deduplicar_roles(cache_roles[itwin_id] + roles_membros)

                role = resolver_role(linha.role, cache_roles[itwin_id])
                usuario = localizar_usuario_por_email(cache_usuarios[itwin_id], linha.email)
                role_ids_atuais = extrair_role_ids_usuario(usuario or {})
                role_ids_desejados = list(dict.fromkeys(role_ids_atuais + [role["id"]]))
                status = "participante" if usuario else "nao_encontrado"
                acao = "atualizar_roles" if usuario else "adicionar_usuario"
                if usuario and role["id"] in role_ids_atuais:
                    acao = "sem_alteracao"

                operacoes.append(
                    {
                        "email": linha.email,
                        "linhaPlanilha": linha.linha,
                        "projetoInformado": linha.projeto,
                        "nomeProjeto": projeto["nome"],
                        "origemProjectWise": projeto.get("origemProjectWise"),
                        "criterioCruzamento": projeto.get("criterioCruzamento"),
                        "connectSpaceId": projeto["connectSpaceId"],
                        "projectId": itwin_id,
                        "acaoSolicitada": "incluir_rbac",
                        "acaoEfetiva": acao,
                        "status": status,
                        "membro": usuario or {},
                        "roleInformada": linha.role,
                        "roleResolvida": role,
                        "roleIdsDesejados": role_ids_desejados,
                        "rolesDisponiveis": [
                            {"id": item["id"], "displayName": item["displayName"], "type": item["type"]}
                            for item in cache_roles[itwin_id]
                        ],
                    }
                )
            except Exception as erro:
                print(f"  [ERRO] Falha ao preparar {linha.email} em {projeto.get('nome')}: {erro}")
                operacoes.append(
                    {
                        "email": linha.email,
                        "linhaPlanilha": linha.linha,
                        "projetoInformado": linha.projeto,
                        "nomeProjeto": projeto.get("nome"),
                        "origemProjectWise": projeto.get("origemProjectWise"),
                        "criterioCruzamento": projeto.get("criterioCruzamento"),
                        "connectSpaceId": projeto.get("connectSpaceId"),
                        "projectId": itwin_id,
                        "acaoSolicitada": "incluir_rbac",
                        "acaoEfetiva": "erro_preparacao_rbac",
                        "status": "erro_preparacao_rbac",
                        "membro": {},
                        "roleInformada": linha.role,
                        "erro": str(erro),
                    }
                )

    return operacoes


def exibir_previa_rbac(operacoes: list[dict[str, Any]]) -> None:
    print("\n[8/9] Previa consolidada RBAC:")
    aplicaveis = [op for op in operacoes if op.get("acaoEfetiva") in {"adicionar_usuario", "atualizar_roles"}]
    usuarios = sorted({str(op.get("email") or "") for op in operacoes if op.get("email")})
    projetos = sorted({str(op.get("projectId") or "") for op in operacoes if op.get("projectId")})
    print(f"- Usuarios: {len(usuarios)}")
    print(f"- Projetos: {len(projetos)}")
    print(f"- Operacoes planejadas: {len(operacoes)}")
    print(f"- Operacoes aplicaveis: {len(aplicaveis)}")

    for indice, op in enumerate(operacoes, start=1):
        role = op.get("roleResolvida") or {}
        print(
            f"{indice:03d}. Linha {op.get('linhaPlanilha') or '-'} | {op.get('email')} | "
            f"{op.get('nomeProjeto') or op.get('projectId')} | {op.get('acaoEfetiva')}"
        )
        if role:
            print(f"     Role: {role.get('displayName')} [{role.get('type')}]")
        if op.get("erro"):
            print(f"     Erro: {op.get('erro')}")


def confirmar_aplicacao_rbac(operacoes: list[dict[str, Any]]) -> bool:
    aplicaveis = [
        op for op in operacoes if op.get("acaoEfetiva") in {"adicionar_usuario", "atualizar_roles"}
    ]
    ignoradas = len(operacoes) - len(aplicaveis)

    if not aplicaveis:
        print("\n[INFO] Nenhuma operacao aplicavel foi encontrada para executar no RBAC.")
        return False

    print("\nConfirmacao final")
    print(f"- Operacoes que alteram o RBAC: {len(aplicaveis)}")
    if ignoradas:
        print(f"- Operacoes apenas informativas/ignoradas: {ignoradas}")

    while True:
        resposta = input("Aplicar as operacoes no Bentley RBAC? [S/N]: ").strip().lower()
        if resposta in {"s", "sim", "y", "yes"}:
            return True
        if resposta in {"n", "nao", "não", "no"}:
            return False
        print("Informe S para sim ou N para nao.")


def aplicar_operacao_rbac(page: Page, op: dict[str, Any]) -> dict[str, Any]:
    acao = op.get("acaoEfetiva")
    itwin_id = op["projectId"]
    email = op["email"]

    if acao == "sem_alteracao":
        return {"status": "sem_alteracao", "mensagem": "Usuario ja possui a role desejada."}
    if acao == "erro_preparacao_rbac":
        return {"status": "ignorado_erro_preparacao", "mensagem": op.get("erro")}
    if acao == "adicionar_usuario":
        payload = {
            "members": [{"email": email, "roleIds": [op["roleResolvida"]["id"]]}],
            "customMessage": "",
        }
        body = exigir_rbac_ok(
            rbac_api(page, "POST", f"/{itwin_id}/members/users", payload),
            f"Adicionar usuario {email} em {itwin_id}",
        )
        return {"status": "adicionado", "body": body}
    if acao == "atualizar_roles":
        usuario = op.get("membro") or {}
        user_id = extrair_id_usuario(usuario)
        if not user_id:
            raise RuntimeError(f"Usuario existente sem id para atualizar: {email}")
        payload = {"roleIds": op["roleIdsDesejados"]}
        body = exigir_rbac_ok(
            rbac_api(page, "PATCH", f"/{itwin_id}/members/users/{user_id}", payload),
            f"Atualizar roles de {email} em {itwin_id}",
        )
        return {"status": "roles_atualizadas", "body": body}

    return {"status": "ignorado", "mensagem": f"Acao nao aplicavel: {acao}"}


def aplicar_operacoes_rbac(page: Page, operacoes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    print("\n[9/9] Aplicando alteracoes RBAC confirmadas...")
    resultados = []
    for indice, op in enumerate(operacoes, start=1):
        print(
            f"- Operacao {indice}/{len(operacoes)}: "
            f"{op.get('email')} | {op.get('nomeProjeto') or op.get('projectId')} [{op.get('acaoEfetiva')}]"
        )
        try:
            resultado = aplicar_operacao_rbac(page, op)
            print(f"  [OK] {resultado['status']}")
        except Exception as erro:
            resultado = {"status": "erro", "erro": str(erro)}
            print(f"  [ERRO] {erro}")

        resultados.append(
            {
                "projectId": op.get("projectId"),
                "nomeProjeto": op.get("nomeProjeto"),
                "origemProjectWise": op.get("origemProjectWise"),
                "criterioCruzamento": op.get("criterioCruzamento"),
                "email": op.get("email"),
                "acao_solicitada": op.get("acaoSolicitada"),
                "acao_planejada": op.get("acaoEfetiva") or op.get("status"),
                "situacao_usuario": op.get("status"),
                "resultado": resultado,
            }
        )
    return resultados


def salvar_log_erro_execucao(erro: Exception, etapa: str, arquivo_projetos_execucao: Optional[Path]) -> None:
    PASTA_LOGS.mkdir(parents=True, exist_ok=True)
    arquivo = PASTA_LOGS / f"rbac_inclusao_roles_erro_{nome_execucao()}.json"
    dados = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "etapa": etapa,
        "erro": str(erro),
        "tipoErro": type(erro).__name__,
        "arquivoProjetosExecucao": str(arquivo_projetos_execucao.resolve()) if arquivo_projetos_execucao else "",
        "traceback": traceback.format_exc(),
    }
    with arquivo.open("w", encoding="utf-8") as saida:
        json.dump(dados, saida, indent=2, ensure_ascii=False)
    print(f"[OK] Log de erro: {arquivo.resolve()}")


def apagar_json_temporario(arquivo: Optional[Path], preservar: bool) -> None:
    if preservar or not arquivo:
        return
    try:
        if arquivo.exists():
            arquivo.unlink()
            print(f"[INFO] JSON temporario removido: {arquivo.resolve()}")
    except Exception as erro:
        print(f"[AVISO] Nao foi possivel remover JSON temporario {arquivo}: {erro}")


def main() -> None:
    preparar_pasta_logs()
    arquivo_projetos_execucao: Optional[Path] = None
    preservar_json_projetos = False
    etapa = "inicio"

    with sync_playwright() as playwright:
        browser: Optional[Browser] = None
        try:
            etapa = "selecionar_projetos_projectwise"
            arquivo_projetos_execucao = executar_seletor_projectwise()
            preservar_json_projetos = True

            etapa = "carregar_projetos_para_rbac"
            projetos = carregar_projetos_para_fluxo(arquivo_projetos_execucao)

            etapa = "coletar_usuarios"
            modo_usuarios = solicitar_modo_usuarios()
            linhas = carregar_linhas_planilha_interativa() if modo_usuarios == "lote_xlsx" else carregar_linhas_usuario_unico()

            etapa = "abrir_navegador"
            browser, _context, page = iniciar_navegador(playwright)
            etapa = "abrir_portal_rbac"
            abrir_portal_rbac(page, projetos[0]["projectId"])

            etapa = "selecionar_role_rbac"
            linhas = preencher_roles_pendentes(page, projetos[0]["projectId"], linhas)

            etapa = "montar_operacoes"
            operacoes = montar_operacoes_rbac(page, projetos, linhas)
            exibir_previa_rbac(operacoes)

            if confirmar_aplicacao_rbac(operacoes):
                etapa = "aplicar_operacoes"
                resultados = aplicar_operacoes_rbac(page, operacoes)
            else:
                print("\n[OK] Execucao cancelada pelo usuario. Nenhuma alteracao foi aplicada.")
                resultados = [
                    {
                        "projectId": op.get("projectId"),
                        "nomeProjeto": op.get("nomeProjeto"),
                        "origemProjectWise": op.get("origemProjectWise"),
                        "criterioCruzamento": op.get("criterioCruzamento"),
                        "email": op.get("email"),
                        "acao_solicitada": op.get("acaoSolicitada"),
                        "acao_planejada": op.get("acaoEfetiva") or op.get("status"),
                        "situacao_usuario": op.get("status"),
                        "resultado": {
                            "status": "cancelado_pelo_usuario",
                            "mensagem": "Usuario nao confirmou a aplicacao final.",
                        },
                    }
                    for op in operacoes
                ]

            etapa = "salvar_log_final"
            salvar_log(operacoes, resultados)
            preservar_json_projetos = False
            print("\nFinalizado.")
            input("Pressione ENTER para fechar o navegador...")

        except KeyboardInterrupt:
            print("\nExecucao interrompida pelo usuario.")
        except Exception as erro:
            print(f"\n[ERRO] {erro}")
            salvar_log_erro_execucao(erro, etapa, arquivo_projetos_execucao)
            if arquivo_projetos_execucao and arquivo_projetos_execucao.exists():
                print(f"[INFO] JSON de projetos preservado para diagnostico: {arquivo_projetos_execucao.resolve()}")
            input("Pressione ENTER para fechar o navegador...")
        finally:
            if browser:
                browser.close()
            apagar_json_temporario(arquivo_projetos_execucao, preservar_json_projetos)


if __name__ == "__main__":
    main()
